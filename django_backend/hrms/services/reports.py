<<<<<<< HEAD
from ..models import Attendance, Employee
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
import io


def attendance_excel(employee_id, month):
    """
    Generate Excel report for employee attendance.
    month format: YYYY-MM
    """
    try:
        employee = Employee.objects.get(id=employee_id)
        year, month_num = map(int, month.split('-'))
        
        attendances = Attendance.objects.filter(
            employee=employee,
            date__year=year,
            date__month=month_num
        ).order_by('date')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Headers
        ws['A1'] = 'Date'
        ws['B1'] = 'Check In'
        ws['C1'] = 'Check Out'
        ws['D1'] = 'Status'
        ws['E1'] = 'Location'
        
        # Data
        row = 2
        for attendance in attendances:
            ws[f'A{row}'] = attendance.date
            ws[f'B{row}'] = attendance.check_in.strftime('%H:%M:%S') if attendance.check_in else ''
            ws[f'C{row}'] = attendance.check_out.strftime('%H:%M:%S') if attendance.check_out else ''
            ws[f'D{row}'] = attendance.status
            ws[f'E{row}'] = attendance.location
            row += 1
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=attendance_{employee.name}_{month}.xlsx'
        return response
        
    except Employee.DoesNotExist:
        return None
    except Exception as e:
        return None
=======
from io import BytesIO

import openpyxl
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


def attendance_excel(rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(["Employee", "Date", "Check In", "Check Out", "Status", "Late", "Work", "Overtime"])
    for row in rows:
        sheet.append([
            row.employee.name,
            row.date.isoformat(),
            row.check_in.strftime("%H:%M") if row.check_in else "",
            row.check_out.strftime("%H:%M") if row.check_out else "",
            row.status,
            row.late_minutes,
            row.work_minutes,
            row.overtime_minutes,
        ])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def payslip_pdf(slip):
    stream = BytesIO()
    pdf = canvas.Canvas(stream, pagesize=A4)
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(60, 780, "AI HRMS Payslip")
    pdf.setFont("Helvetica", 11)
    lines = [
        f"Employee: {slip.employee.name}",
        f"Code: {slip.employee.employee_code}",
        f"Month: {slip.month}",
        f"Paid days: {slip.paid_days}",
        f"Absent days: {slip.absent_days}",
        f"Overtime minutes: {slip.overtime_minutes}",
        f"Gross salary: {slip.gross_salary}",
        f"Deductions: {slip.deductions}",
        f"Net salary: {slip.net_salary}",
    ]
    y = 735
    for line in lines:
        pdf.drawString(60, y, line)
        y -= 24
    pdf.showPage()
    pdf.save()
    stream.seek(0)
    return ContentFile(stream.read(), name=f"payslip-{slip.employee.employee_code}-{slip.month}.pdf")
>>>>>>> 38eaefb (Production ready: Django backend configured for Render deployment, frontend deployed to Vercel)
