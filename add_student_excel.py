from openpyxl import Workbook, load_workbook
import os

file_name = "students.xlsx"

def create_file():
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["ID", "Name"])
        wb.save(file_name)

def add_student():
    id = input("Enter ID: ")
    name = input("Enter Name: ")

    create_file()

    wb = load_workbook(file_name)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == id:
            print("Student already exists ❌")
            return

    ws.append([id, name])
    wb.save(file_name)

    print("Student Added Successfully ✅")

add_student()